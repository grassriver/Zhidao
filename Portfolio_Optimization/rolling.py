# -*- coding: utf-8 -*-
"""
Rolling Optimization

Created on Sat Apr  7 16:06:00 2018

@author: Kai Zheng
"""

import numpy as np
import pandas as pd
import general_tools as tool
import sqlite3 as sql
import stock_screener as sc
from portfolio_class import Portfolio
from stock_class import Stock
import Portfolio_Optimization.mv as mv
import max_drawdown as md
import utils
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
#import Portfolio_Optimization.analytical as ant
#import CAPM.capm as capm

# =============================================================================
# input: stock_list
# output: portfolio balance
# =============================================================================

#%%
# forward fill
def rolling(conn, stock_list, start, end, step='M', cap=1000000, 
            backfill=False, mu_method=None, sigma_method=None, opt_method=None,
            rf=0.03, outpath='./Portfolio_Optimization/output/', outfile='sample.xlsx'):
    delta = pd.Timedelta(1, unit=step)
    start = pd.to_datetime(start)
    end = pd.to_datetime(end)
    now = start
    stocks_price_old = tool.get_stocks_price_old(conn, stock_list)
    business_calendar = tool.get_business_calendar(conn)
    industry = tool.get_industry(conn, stock_list)
    benchmark_old = tool.get_benchmark_old(conn)
    cap_old = cap
    
#    wb = Workbook()
    try:
        wb = openpyxl.load_workbook(outpath+outfile)
    except:
        wb = Workbook()
    sheetname = '-'.join([mu_method.name, sigma_method.name, opt_method.name])
    try:
        ws = wb.get_sheet_by_name(sheetname)
    except:
        ws = wb.create_sheet(sheetname)
    
    if mu_method is not None:
        mu_method.update(stocks_price_old=stocks_price_old,
                         business_calendar=business_calendar,
                         industry=industry,
                         backfill=backfill)
    if sigma_method is not None:
        sigma_method.update(stocks_price_old=stocks_price_old,
                            business_calendar=business_calendar,
                            industry=industry,
                            backfill=backfill)
    balance = pd.Series()
    while now < end:
        now_end = now + delta
        (balance, 
         cap, 
         mk_port, 
         weights,
         mu, 
         sigma, 
         stocks) = single_stage_opt(
                         conn, stock_list, now, 
                         now_end, cap, balance, backfill,rf,
                         mu_method=mu_method,
                         sigma_method=sigma_method,
                         opt_method=opt_method,
                         stocks_price_old=stocks_price_old,
                         business_calendar=business_calendar,
                         industry=industry,
                         benchmark_old=benchmark_old
                         )
        if sigma_method.name == 'opt_quadratic_risky_restricted':
            sigma_method.update(w0=weights)
        now = now + delta
        if now + delta > end:
            delta = end - now
        opt_info = opt_info_output(stock_list, weights, mu, sigma, stocks, thres=0.05)
        utils.write_to_excel(opt_info, ws)
    
    
    combined_balance = combine_balances(conn, start, end, stock_list, 
                                        backfill, balance, cap_old, benchmark='sh000016')
    plot_xlsx(combined_balance, outpath, sheetname, ws)
    wb.save(outpath+outfile)
    wb.close()
    
    return balance

def single_stage_opt(conn, stock_list, start, end, cap, balance, backfill, rf, 
                     mu_method=None, sigma_method=None, opt_method=None, 
                     stocks_price_old=None, business_calendar=None, 
                     industry=None, benchmark_old=None):
    #================ check date availability===========================
    start = pd.to_datetime(
            tool.first_trading_day(conn, start.strftime('%Y-%m-%d'), 
                                         business_calendar=business_calendar))
    #================ predict mean anv cov of returns===================

    mu, mu_method = update_mu(
            mu_method, conn, stock_list, start, backfill, 
            stocks_price_old, business_calendar, industry)
 
    sigma, sigma_method = update_sigma(
            sigma_method, conn, stock_list, start, backfill, 
            stocks_price_old, business_calendar, industry)
        
    weights, opt_method = update_opt(opt_method, mu, sigma)
        
    balance, mk_port, stocks = update_balance(
            balance, conn, stock_list, start, end, 
            backfill, stocks_price_old, business_calendar, 
            industry, weights, cap, benchmark_old, rf)

    cap = balance[-1]
    
    return balance, cap, mk_port, weights, mu, sigma, stocks

def opt_info_output(code_list, weights, mu, sigma, stocks, thres=0.05):
    stocks_selected = np.array(code_list)[weights>thres].tolist()
    proj_ret = pd.Series(mu[stocks_selected], name='proj_ret')
    temp = np.in1d(sigma.columns, stocks_selected)
    proj_cov = sigma.loc[temp,temp]
    proj_cov = proj_cov.set_index(proj_cov.columns)
    proj_std = pd.Series(np.sqrt(np.diag(proj_cov)), 
                         index=proj_cov.columns, 
                         name='proj_std')
    rlz_daily_returns = stocks.daily_returns
    rlz_ret = rlz_daily_returns.mean()[stocks_selected]*252
    rlz_ret.name='rlz_ret'
    temp = np.in1d(rlz_daily_returns.columns, stocks_selected)
    rlz_cov = rlz_daily_returns.cov().loc[temp, temp]*252
    rlz_std = pd.Series(np.sqrt(np.diag(rlz_cov)), 
                        index=rlz_cov.columns, 
                        name='rlz_std')
    
    df = pd.DataFrame({'weights':weights}, index=code_list)
    df = df[np.in1d(df.index,stocks_selected)]
    df = df.join(proj_ret)
    df = df.join(proj_std)
    df = df.join(rlz_ret)
    df = df.join(rlz_std)
    
    return df

def drawdown_table(balance, top=5):
    ret = np.log(balance/balance.shift(1))
    df = md.gen_drawdown_table(ret, top=top)
    return df

def drawdown_plot(balance, top=5):
    ret = np.log(balance/balance.shift(1))
    ax = md.plot_drawdown_periods(ret, top=top)
    return ax

def gen_ewet_port_balance(conn, start, end, code_list, capital, backfill):
    # equal weight portfolio
    ewet = tool.portfolio_construct_by_weight(
                conn, start.strftime('%Y-%m-%d'), code_list,
                capital=capital, backfill=True)
    ewet_port = Portfolio(conn, ewet, start=start, end=end, backfill=True)
    ewet_port_balance = ewet_port.port_daily_balance()
    return ewet_port_balance

def benchmark_price(conn, benchmark):
    # benchmark performance #hs50
    query = 'select date, close from index_price where code = "' + benchmark + '"'
    bench_price = pd.read_sql(query, conn)
    bench_price['date'] = pd.to_datetime(bench_price['date'])
    bench_price.set_index('date', inplace=True)
    return bench_price

def combine_balances(conn, start, end, code_list, backfill, balance, capital, benchmark='sh000016'):
    # combine together
    bench_price = benchmark_price(conn, benchmark)
    ewet_port_balance = gen_ewet_port_balance(conn, start, end, code_list, capital, backfill)
    balance = capital/balance[0]*balance
    ewet_port_balance = capital/ewet_port_balance[0]*ewet_port_balance
    
    df = pd.merge(pd.DataFrame(balance), bench_price,
                  left_index=True, right_index=True, how = 'left')
    df = df.join(pd.DataFrame({'ewet':ewet_port_balance}))
    df.columns=['portfolio', 'benchmark', 'equal weight portfolio']
    df['benchmark'] = capital/df['benchmark'][0]*df['benchmark']
    return df

def plot_combined_balance(combined_balance):
    return combined_balance.plot(figsize=(16,9))

def update_mu(mu_method, conn, stock_list, start, backfill, 
              stocks_price_old, business_calendar, industry):
    # estimate mean of returns
    if mu_method is None:
        mu = mv.hist_expect_mu(
                    conn, stock_list,
                    start=(start-pd.Timedelta(30, unit='d')).strftime('%Y-%m-%d'),
                    end=(start).strftime('%Y-%m-%d'),
                    backfill = backfill, 
                    stocks_price_old=stocks_price_old,
                    business_calendar=business_calendar,
                    industry=industry)
    elif mu_method.name == 'capm':
        mu_method.update(start=start.strftime('%Y-%m-%d'))
        mu = mu_method.run()
    elif mu_method.name == 'hist':
        lookback_win = mu_method.kwargs['lookback_win']
        mu_method.update(start = (start-pd.Timedelta(lookback_win, unit='d')).strftime('%Y-%m-%d'),
                         end = (start-pd.Timedelta(1)).strftime('%Y-%m-%d'))
        mu = mu_method.run()
    else:
        raise ValueError('mu_method not right!')
        
    return mu, mu_method

def update_sigma(sigma_method, conn, stock_list, start, backfill, 
              stocks_price_old, business_calendar, industry):
    # estimate covariance matrix of returns
    if sigma_method is None:
        sigma = mv.hist_expect_sigma(
                    conn, stock_list,
                    start=(start-pd.Timedelta(30, unit='d')).strftime('%Y-%m-%d'),
                    end=(start).strftime('%Y-%m-%d'),
                    backfill = backfill, 
                    stocks_price_old=stocks_price_old,
                    business_calendar=business_calendar,
                    industry=industry)
    elif sigma_method.name == 'hist':
        lookback_win = sigma_method.kwargs['lookback_win']
        sigma_method.update(
                start = (start-pd.Timedelta(lookback_win, unit='d')).strftime('%Y-%m-%d'),
                end = (start-pd.Timedelta(1)).strftime('%Y-%m-%d'))
        sigma = sigma_method.run()
    elif sigma_method.name == 'barra':
        sigma_method.update(
                start = (start-pd.Timedelta(1, unit='d')).strftime('%Y-%m-%d'))
        sigma = sigma_method.run()
    else:
        raise ValueError('sigma_method not right!')
    
    return sigma, sigma_method

def update_opt(opt_method, mu, sigma):
    #================= mean-variance optimization ======================
    if opt_method is None:
        # use numeric solution
        opts, optv= mv.opt_s_v(mu,sigma)
        weights = opts.x
    elif opt_method.name in ['opt_quadratic', 'opt_quadratic_risky', 'opt_s', 'opt_v']:
        opt_method.update(mu=mu, sigma=sigma)
        weights = opt_method.run()
    elif opt_method.name == 'opt_quadratic_risky_restricted':
        opt_method.update(mu=mu, sigma=sigma)
        weights = opt_method.run()
    elif opt_method.name.split('-')[0] in ['opt_quadratic', 'opt_quadratic_risky', 'opt_s', 'opt_v']:
        opt_method.update(mu=mu, sigma=sigma)
        weights = opt_method.run()
    else:
        raise ValueError('opt_method not right!')
    
    if np.isnan(weights).any():
        print(mu)
        print(sigma)
        raise ValueError('optimal weights error')
    print(weights)
        
    return weights, opt_method


    
def update_balance(balance, conn, stock_list, start, end, 
                   backfill, stocks_price_old, business_calendar, 
                   industry, weights, cap, benchmark_old, rf):
    #================== backtest ==========================
    
    # calculate shares of each stock according to weights
    if backfill==False:
        mport = tool.portfolio_construct_by_weight(conn, 
                                                   start.strftime('%Y-%m-%d'), 
                                                   pd.Series(stock_list), 
                                                   weights = weights,
                                                   capital = cap)
    else:
        stocks = Stock(conn, stock_list, 
                       start=start.strftime('%Y-%m-%d'), 
                       end=end.strftime('%Y-%m-%d'),
                       backfill = backfill, 
                       stocks_price_old=stocks_price_old,
                       business_calendar=business_calendar,
                       industry=industry)
        date = stocks.start
        stock_price = stocks.price[date].reset_index()
        shares = np.floor(cap*weights/stock_price['close']/100)*100 
        mport = pd.DataFrame({'code': stock_list, 'shares': shares})
    
    # construct a portfolio according to mport
    mk_port = Portfolio(
                    conn, mport, 
                    start=start.strftime('%Y-%m-%d'), 
                    end=end.strftime('%Y-%m-%d'), 
                    backfill=backfill,
                    stocks_price_old = stocks_price_old,
                    business_calendar=business_calendar,
                    industry=industry,
                    benchmark_old=benchmark_old)
    # update balance
    daily_balance =  mk_port.port_daily_balance()
    if (1-weights.sum())>1e-6:
        cash_cap = cap-daily_balance[0]
        cash_balance = [cash_cap*np.exp(rf/252*i) 
                            for i in range(1, len(daily_balance)+1)]
        cash_balance = pd.Series(cash_balance, index=daily_balance.index)
        daily_balance = daily_balance + cash_balance
        print(weights.sum())
    else:
        daily_balance = daily_balance + cap - daily_balance[0]
    
    balance = balance.append(daily_balance)
    
    return balance, mk_port, stocks

def plot_xlsx(combined_balance, outpath, sheetname, ws):
    combined_balance.plot(figsize=(16,9))
    plt.savefig(outpath+sheetname)
    img = Image(outpath+sheetname+'.png')
    utils.insert_image(img, ws)


if __name__ == '__main__':
    conn_path = 'D:/Kaizheng/Working_directory/portfolio_intelligence/PI/Data/data.db'
    conn = sql.connect(conn_path)
    date = '2015-10-31'
    start='2016-01-05'
    end='2016-10-31'

    stock_list = sc.stock_screener_ranking(conn_path, date=date, 
                                           var_list=['net_profit_ratio','roe','eps'],
                                           rank_by = 'roe',order = 'ascending',top=10)
    stocks = tool.portfolio_construct(conn,start = '2016-01-05',code_list=stock_list.Code,
                                      construct_type='weight',equal=True)
    code_list = list(stocks.code)

    # Plot Efficient Frontier
    mu,sigma = mv.hist_expect(conn, code_list, start=start, end=end) #cheating here
    mv.plot_eff_fter(mu, sigma)
    weights = mv.tan_port(mu,sigma).x

    balance = rolling(conn, code_list, start, end)
    print(balance)
    balance.plot()