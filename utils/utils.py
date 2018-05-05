# -*- coding: utf-8 -*-
import numpy as np
from portfolio_class import Portfolio
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

#%%
def check_symmetric(a, tol=1e-8):
    return np.allclose(a, a.T, atol=tol)

def write_to_excel(df, ws):
    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)
        
def insert_image(img, ws):
    ws.add_image(img, 'A1')
    
#%% demo for playing with openpyxl
# =============================================================================
# wb = openpyxl.Workbook()
# wb = openpyxl.load_workbook('sample.xlsx')
# # grab the active worksheet
# ws = wb.active
# 
# # Data can be assigned directly to cells
# ws['A1'] = 42
# 
# # Rows can also be appended
# ws.append([1, 2, 3])
# 
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
# 
# for r in dataframe_to_rows(opt_info, index=True, header=True):
#     ws.append(r)
# 
# for cell in ws['A'] + ws[1]:
#     cell.style = 'Pandas'
# 
# # Save the file
# wb.save("sample.xlsx")
# wb.close()
# =============================================================================
